import csv
import re
from datetime import datetime
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook

from app.database import get_supabase_client
from app.models.siswa import (
    SiswaCreateRequest,
    SiswaDetailResponse,
    SiswaImportResponse,
    SiswaListResponse,
    SiswaUpdateRequest,
)
from app.utils.auth import get_current_user_profile

router = APIRouter()

SISWA_SELECT_FIELDS = (
    "id,ra_id,nama,kelompok_id,status_aktif,nis,"
    "nisn,nik,tempat_lahir,tanggal_lahir,tingkat_rombel,umur_text,"
    "jenis_kelamin,alamat,no_telepon,kebutuhan_khusus,disabilitas,"
    "nomor_kip_pip,nama_ayah_kandung,nama_ibu_kandung,nama_wali"
)

EMIS_HEADER_ALIASES = {
    "nama": ["nama", "nama lengkap"],
    "nisn": ["nisn"],
    "nik": ["nik"],
    "tempat_lahir": ["tempat lahir"],
    "tanggal_lahir": ["tanggal lahir", "tgl lahir"],
    "tingkat_rombel": ["tingkat - rombel", "tingkat rombel", "rombel"],
    "umur_text": ["umur", "usia"],
    "status": ["status", "status aktif", "status_aktif"],
    "jenis_kelamin": ["jenis kelamin", "jk"],
    "alamat": ["alamat"],
    "no_telepon": ["no telepon", "nomor telepon", "telepon", "no hp", "hp"],
    "kebutuhan_khusus": ["kebutuhan khusus"],
    "disabilitas": ["disabilitas"],
    "nomor_kip_pip": ["nomor kip/pip", "nomor kip pip", "kip/pip", "kip pip"],
    "nama_ayah_kandung": ["nama ayah kandung", "nama ayah"],
    "nama_ibu_kandung": ["nama ibu kandung", "nama ibu"],
    "nama_wali": ["nama wali"],
}


def _validate_kelompok_access(supabase, ra_id: str, kelompok_id: str):
    kelompok = (
        supabase.table("kelompok")
        .select("id")
        .eq("id", kelompok_id)
        .eq("ra_id", ra_id)
        .limit(1)
        .execute()
    )
    if not kelompok.data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"kelompok_id {kelompok_id} tidak valid untuk RA ini",
        )


def _normalize_header(value: str) -> str:
    lowered = (value or "").strip().lower()
    lowered = lowered.replace("_", " ")
    lowered = re.sub(r"\s+", " ", lowered)
    return lowered


def _clean_cell(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("'"):
        text = text[1:].strip()
    return text or None


def _to_bool_status(value, default: bool = True) -> bool:
    cleaned = (_clean_cell(value) or "").lower()
    if not cleaned:
        return default
    return cleaned not in {"false", "0", "no", "tidak", "nonaktif", "inactive"}


def _to_date_string(value) -> str | None:
    cleaned = _clean_cell(value)
    if not cleaned:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return cleaned


def _get_nis_base(
    tingkat_rombel: str | None,
    resolved_kelompok_id: str | None,
    kelompok_name_by_id: dict[str, str],
) -> int | None:
    kelompok_ref = _normalize_header(tingkat_rombel or "")
    if not kelompok_ref and resolved_kelompok_id:
        kelompok_ref = _normalize_header(kelompok_name_by_id.get(resolved_kelompok_id, ""))

    if "kelompok a" in kelompok_ref:
        return 250000
    elif "kelompok b" in kelompok_ref:
        return 240000

    return None


def _assign_nis_by_alphabet(imported_rows: list[dict], kelompok_name_by_id: dict[str, str]):
    grouped: dict[int, list[dict]] = {}

    for row in imported_rows:
        base = _get_nis_base(
            row.get("tingkat_rombel"),
            row.get("kelompok_id"),
            kelompok_name_by_id,
        )
        if not base:
            row["nis"] = None
            continue
        grouped.setdefault(base, []).append(row)

    for base, rows in grouped.items():
        sorted_rows = sorted(rows, key=lambda item: _normalize_header(item.get("nama") or ""))
        for index, row in enumerate(sorted_rows, start=1):
            row["nis"] = str(base + index)


def _extract_emis_fields(raw_row: dict) -> dict:
    normalized_row = {
        _normalize_header(str(key)): _clean_cell(value)
        for key, value in raw_row.items()
        if key is not None
    }

    mapped = {}
    for canonical_key, aliases in EMIS_HEADER_ALIASES.items():
        value = None
        for alias in aliases:
            alias_normalized = _normalize_header(alias)
            candidate = normalized_row.get(alias_normalized)
            if candidate:
                value = candidate
                break
        mapped[canonical_key] = value

    return mapped


def _build_kelompok_index(kelompok_data: list[dict]) -> dict[str, str]:
    index = {}
    for item in kelompok_data:
        nama = _clean_cell(item.get("nama_kelompok"))
        item_id = item.get("id")
        if not nama or not item_id:
            continue

        normalized = _normalize_header(nama)
        index[normalized] = item_id

        segments = [segment.strip() for segment in normalized.split("-") if segment.strip()]
        for segment in segments:
            index.setdefault(segment, item_id)

    return index


def _get_or_create_kelompok_id_by_name(
    supabase,
    ra_id: str,
    tingkat_rombel: str | None,
    kelompok_index: dict[str, str],
    kelompok_name_by_id: dict[str, str],
):
    if not tingkat_rombel:
        return None

    normalized = _normalize_header(tingkat_rombel)
    if normalized in kelompok_index:
        return kelompok_index[normalized]

    # Fallback fuzzy match against existing index first
    for key, value in kelompok_index.items():
        if key in normalized or normalized in key:
            return value

    # Create new kelompok when not found
    try:
        create_response = (
            supabase.table("kelompok")
            .insert({"ra_id": ra_id, "nama_kelompok": tingkat_rombel})
            .execute()
        )
        created = create_response.data[0] if create_response.data else None
    except Exception:
        return None

    if not created:
        return None

    created_id = created.get("id")
    created_name = _clean_cell(created.get("nama_kelompok")) or tingkat_rombel
    normalized_created_name = _normalize_header(created_name)
    if created_id:
        kelompok_index[normalized_created_name] = created_id
        kelompok_name_by_id[created_id] = created_name

        segments = [segment.strip() for segment in normalized_created_name.split("-") if segment.strip()]
        for segment in segments:
            kelompok_index.setdefault(segment, created_id)

    return created_id


def _resolve_kelompok_id(tingkat_rombel: str | None, fallback_kelompok_id: str | None, kelompok_index: dict[str, str]):
    if fallback_kelompok_id:
        return fallback_kelompok_id
    if not tingkat_rombel:
        return None

    normalized = _normalize_header(tingkat_rombel)
    if normalized in kelompok_index:
        return kelompok_index[normalized]

    for key, value in kelompok_index.items():
        if key in normalized or normalized in key:
            return value

    return None


@router.get("", response_model=SiswaListResponse)
def list_siswa(
    kelompok_id: str | None = Query(default=None),
    current=Depends(get_current_user_profile),
):
    supabase = get_supabase_client()
    ra_id = current["ra_id"]

    try:
        query = (
            supabase.table("siswa")
            .select(SISWA_SELECT_FIELDS)
            .eq("ra_id", ra_id)
            .order("nama")
        )
        if kelompok_id:
            query = query.eq("kelompok_id", kelompok_id)
        response = query.execute()
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal mengambil data siswa: {exc}",
        ) from exc

    return {"success": True, "data": response.data or []}


@router.post("", response_model=SiswaDetailResponse, status_code=status.HTTP_201_CREATED)
def create_siswa(payload: SiswaCreateRequest, current=Depends(get_current_user_profile)):
    supabase = get_supabase_client()
    ra_id = current["ra_id"]

    try:
        if payload.kelompok_id:
            _validate_kelompok_access(supabase, ra_id, payload.kelompok_id)

        insert_data = payload.model_dump(exclude_none=True)
        insert_data["ra_id"] = ra_id
        response = (
            supabase.table("siswa")
            .insert(insert_data)
            .execute()
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gagal menambah siswa: {exc}",
        ) from exc

    return {
        "success": True,
        "message": "Siswa berhasil ditambahkan",
        "data": response.data[0],
    }


@router.put("/{id}", response_model=SiswaDetailResponse)
def update_siswa(id: str, payload: SiswaUpdateRequest, current=Depends(get_current_user_profile)):
    supabase = get_supabase_client()
    ra_id = current["ra_id"]

    update_data = payload.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tidak ada data yang diubah",
        )

    if "kelompok_id" in update_data and update_data["kelompok_id"]:
        _validate_kelompok_access(supabase, ra_id, update_data["kelompok_id"])

    try:
        response = (
            supabase.table("siswa")
            .update(update_data)
            .eq("id", id)
            .eq("ra_id", ra_id)
            .execute()
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gagal mengubah siswa: {exc}",
        ) from exc

    if not response.data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Siswa tidak ditemukan",
        )

    return {
        "success": True,
        "message": "Data siswa berhasil diubah",
        "data": response.data[0],
    }


@router.delete("/{id}", response_model=SiswaDetailResponse)
def delete_siswa(id: str, current=Depends(get_current_user_profile)):
    supabase = get_supabase_client()
    ra_id = current["ra_id"]

    try:
        response = (
            supabase.table("siswa")
            .update({"status_aktif": False})
            .eq("id", id)
            .eq("ra_id", ra_id)
            .execute()
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gagal menonaktifkan siswa: {exc}",
        ) from exc

    if not response.data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Siswa tidak ditemukan",
        )

    return {
        "success": True,
        "message": "Siswa berhasil dinonaktifkan",
        "data": response.data[0],
    }


@router.post("/import", response_model=SiswaImportResponse)
async def import_siswa(
    kelompok_id: str | None = Query(default=None),
    file: UploadFile = File(...),
    current=Depends(get_current_user_profile),
):
    supabase = get_supabase_client()
    ra_id = current["ra_id"]

    if kelompok_id:
        _validate_kelompok_access(supabase, ra_id, kelompok_id)

    try:
        kelompok_response = (
            supabase.table("kelompok")
            .select("id,nama_kelompok")
            .eq("ra_id", ra_id)
            .execute()
        )
        kelompok_data = kelompok_response.data or []
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gagal mengambil data kelompok: {exc}",
        ) from exc

    kelompok_index = _build_kelompok_index(kelompok_data)
    kelompok_name_by_id = {
        item.get("id"): _clean_cell(item.get("nama_kelompok")) or ""
        for item in kelompok_data
        if item.get("id")
    }

    content = await file.read()
    filename = (file.filename or "").lower()
    imported_rows: list[dict] = []
    skipped_no_kelompok = 0
    created_kelompok_count = 0

    try:
        if filename.endswith(".csv"):
            text_stream = StringIO(content.decode("utf-8-sig"))
            reader = csv.DictReader(text_stream)
            for row in reader:
                mapped = _extract_emis_fields(row)
                nama = mapped.get("nama")
                if not nama:
                    continue

                resolved_kelompok_id = _resolve_kelompok_id(
                    mapped.get("tingkat_rombel"),
                    kelompok_id,
                    kelompok_index,
                )
                if not resolved_kelompok_id and mapped.get("tingkat_rombel"):
                    resolved_kelompok_id = _get_or_create_kelompok_id_by_name(
                        supabase,
                        ra_id,
                        mapped.get("tingkat_rombel"),
                        kelompok_index,
                        kelompok_name_by_id,
                    )
                    if resolved_kelompok_id:
                        created_kelompok_count += 1
                if not resolved_kelompok_id:
                    skipped_no_kelompok += 1
                    continue

                imported_rows.append(
                    {
                        "ra_id": ra_id,
                        "nama": nama,
                        "kelompok_id": resolved_kelompok_id,
                        "status_aktif": _to_bool_status(mapped.get("status"), default=True),
                        "nis": None,
                        "nisn": mapped.get("nisn"),
                        "nik": mapped.get("nik"),
                        "tempat_lahir": mapped.get("tempat_lahir"),
                        "tanggal_lahir": _to_date_string(mapped.get("tanggal_lahir")),
                        "tingkat_rombel": mapped.get("tingkat_rombel"),
                        "umur_text": mapped.get("umur_text"),
                        "jenis_kelamin": mapped.get("jenis_kelamin"),
                        "alamat": mapped.get("alamat"),
                        "no_telepon": mapped.get("no_telepon"),
                        "kebutuhan_khusus": mapped.get("kebutuhan_khusus"),
                        "disabilitas": mapped.get("disabilitas"),
                        "nomor_kip_pip": mapped.get("nomor_kip_pip"),
                        "nama_ayah_kandung": mapped.get("nama_ayah_kandung"),
                        "nama_ibu_kandung": mapped.get("nama_ibu_kandung"),
                        "nama_wali": mapped.get("nama_wali"),
                    }
                )
        elif filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="File Excel kosong",
                )

            raw_headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            normalized_headers = [_normalize_header(header) for header in raw_headers]

            if "nama" not in normalized_headers and "nama lengkap" not in normalized_headers:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Header 'Nama' atau 'Nama Lengkap' wajib ada pada file Excel",
                )

            for row_values in rows[1:]:
                if not any(cell is not None and str(cell).strip() for cell in row_values):
                    continue

                row = {}
                for idx, header in enumerate(raw_headers):
                    row[header] = row_values[idx] if idx < len(row_values) else None

                mapped = _extract_emis_fields(row)
                nama = mapped.get("nama")
                if not nama:
                    continue

                resolved_kelompok_id = _resolve_kelompok_id(
                    mapped.get("tingkat_rombel"),
                    kelompok_id,
                    kelompok_index,
                )
                if not resolved_kelompok_id and mapped.get("tingkat_rombel"):
                    resolved_kelompok_id = _get_or_create_kelompok_id_by_name(
                        supabase,
                        ra_id,
                        mapped.get("tingkat_rombel"),
                        kelompok_index,
                        kelompok_name_by_id,
                    )
                    if resolved_kelompok_id:
                        created_kelompok_count += 1
                if not resolved_kelompok_id:
                    skipped_no_kelompok += 1
                    continue

                imported_rows.append(
                    {
                        "ra_id": ra_id,
                        "nama": nama,
                        "kelompok_id": resolved_kelompok_id,
                        "status_aktif": _to_bool_status(mapped.get("status"), default=True),
                        "nis": None,
                        "nisn": mapped.get("nisn"),
                        "nik": mapped.get("nik"),
                        "tempat_lahir": mapped.get("tempat_lahir"),
                        "tanggal_lahir": _to_date_string(mapped.get("tanggal_lahir")),
                        "tingkat_rombel": mapped.get("tingkat_rombel"),
                        "umur_text": mapped.get("umur_text"),
                        "jenis_kelamin": mapped.get("jenis_kelamin"),
                        "alamat": mapped.get("alamat"),
                        "no_telepon": mapped.get("no_telepon"),
                        "kebutuhan_khusus": mapped.get("kebutuhan_khusus"),
                        "disabilitas": mapped.get("disabilitas"),
                        "nomor_kip_pip": mapped.get("nomor_kip_pip"),
                        "nama_ayah_kandung": mapped.get("nama_ayah_kandung"),
                        "nama_ibu_kandung": mapped.get("nama_ibu_kandung"),
                        "nama_wali": mapped.get("nama_wali"),
                    }
                )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Format file tidak didukung. Gunakan CSV atau XLSX",
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gagal parse file impor: {exc}",
        ) from exc

    if not imported_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tidak ada data siswa valid untuk diimpor (cek nama kolom atau mapping kelompok)",
        )

    _assign_nis_by_alphabet(imported_rows, kelompok_name_by_id)

    try:
        supabase.table("siswa").insert(imported_rows).execute()
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Gagal menyimpan data impor siswa: {exc}",
        ) from exc

    return {
        "success": True,
        "message": (
            f"Import siswa berhasil. "
            f"Kelompok baru dibuat: {created_kelompok_count}. "
            f"{skipped_no_kelompok} baris dilewati karena kelompok tidak ditemukan"
        ),
        "imported_count": len(imported_rows),
    }
